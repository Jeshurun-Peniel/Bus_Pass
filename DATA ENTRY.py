from tkinter import *
from tkinter import font
from PIL import ImageTk, Image 
import time
import PySimpleGUI as sg
import openpyxl
import os

create = False
error = 0
header = ["School Name", "School Code", "Student Name", "Class", "Section", 
            "Door Number", "Street Nmae", "Area", "City", "Pincode", "Journey From", "Journey To", "Km", "Fare"]
sg.theme('DarkTeal9')
sg.set_options(font=('Helvetica'), text_color='white')

desktop = os.path.join(os.path.join(os.environ["HOMEPATH"]),"Desktop")

Main_layout = [   
    [sg.Text('Student Name:', size=(12,1)), sg.Push(), sg.InputText(size=(3,1),key='Student Initial'),sg.InputText(size=(40,1),key='Student Name'),],
    [sg.Text('Class', size=(12,1)), sg.InputCombo(('VI', 'VII','VIII','IX','X','XI','XII'), size=(4, 1),key='Class'), 
    sg.Text('Sec'),sg.InputCombo(('A', 'A1','A2', 'A/A','B','B1','C','D','D1', 'D2','E','F','G','GM','H'), size=(4, 1),key='Section')],
    [sg.Text('Address:')],
    [sg.Text('Door Number', size=(12,1)), sg.InputText(size=(10,1),key='Door Number')], 
    [sg.Text('Street Name',size=(12,1)), sg.InputText(key='Street Name')],
    [sg.Text("Area", size=(12,1)), sg.InputText(key='Area')], 
    [sg.Text('City', size=(12,1)), sg.InputCombo(('TIRUNELVELI', 'THOOTHUKUDI'), size=(12, 1),key='City'), 
     sg.Text('Pincode'), sg.Input(size=(10,1),key='Pincode')],
    [sg.Text('Journey From        '),sg.Text("JOURNEY FROM WILL BE THE SAME AS AREA !",text_color='white', background_color='black'), sg.InputText("DON'T TYPE HERE", key='Journey from',visible=False)], 
    [sg.Text('Journey To', size=(12,1)), sg.InputCombo(('PALAI CHRISTHURAJA', 'PALAI WATER TANK', 'SAMATHANAPURAM'), size=(18, 1),key='Journey to')],
    [sg.Text("Km", size=(12,1)), sg.Input(size=(10,1), key='Km')],
    [sg.Text("Fare", size=(12,1)), sg.InputText(size=(10,1),key='Fare')],
    [sg.Submit(), sg.Button('Clear'),sg.Button('Reset',button_color='red'), sg.Exit()]   
]
def start_up():
    w=Tk()
#Using piece of code from old splash screen
    width_of_window = 427
    height_of_window = 250
    screen_width = w.winfo_screenwidth()
    screen_height = w.winfo_screenheight()
    x_coordinate = (screen_width/2)-(width_of_window/2)
    y_coordinate = (screen_height/2)-(height_of_window/2)
    w.geometry("%dx%d+%d+%d" %(width_of_window,height_of_window,x_coordinate,y_coordinate))
#w.configure(bg='#ED1B76')
    w.overrideredirect(1) #for hiding titlebar

    Frame(w, width=427, height=250, bg='#272727').place(x=0,y=0)
    label1=Label(w, text='BUSPASS ENTRY', fg='white', bg='#272727') #decorate it 
    label1.configure(font=("Game Of Squids", 24, "bold"))   #You need to install this font in your PC or try another one
    label1.place(x=80,y=90)


    label2=Label(w, text='Loading...', fg='white', bg='#272727') #decorate it 
    label2.configure(font=("Calibri", 11))
    label2.place(x=10,y=215)

#making animation

    image_a=ImageTk.PhotoImage(Image.open('loading2.png'))
    image_b=ImageTk.PhotoImage(Image.open('loading1.png'))

    for i in range(5): #5loops
        l1=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=180, y=145)
        l2=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=200, y=145)
        l3=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=220, y=145)
        l4=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=240, y=145)
        w.update_idletasks()
        time.sleep(0.5)

        l1=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=180, y=145)
        l2=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=200, y=145)
        l3=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=220, y=145)
        l4=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=240, y=145)
        w.update_idletasks()
        time.sleep(0.5)

        l1=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=180, y=145)
        l2=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=200, y=145)
        l3=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=220, y=145)
        l4=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=240, y=145)
        w.update_idletasks()
        time.sleep(0.5)

        l1=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=180, y=145)
        l2=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=200, y=145)
        l3=Label(w, image=image_b, border=0, relief=SUNKEN).place(x=220, y=145)
        l4=Label(w, image=image_a, border=0, relief=SUNKEN).place(x=240, y=145)
        w.update_idletasks()
        time.sleep(0.5)

    w.destroy()

def get_file():

    layout3 = [[sg.Text("SELECT A EXCEL FILE OR CREATE A NEW EXCEL FILE...", font=40)],
            [sg.Button('Open'),sg.Button('Create'),sg.Button('Exit',button_color='red')]]
    layout4 = [[sg.Text("CREATE A NEW EXCEL FILE", font=40)],
        [sg.Text("Enter the File Name: ", font=40),sg.InputText(key='name')],
            [sg.Button('Save'),sg.Button('Cancel')]]

    window = sg.Window("BUS PASS ENTRY", layout3,icon='down.ico')
    window2= sg.Window("BUS PASS ENTRY", layout4,icon='create.ico')

    while True:
        global error
        event2,values2 = window.read()
        if event2 == "Exit" or event2 == sg.WIN_CLOSED:
            error = 1
            window.close()
            break
        elif event2 == "Create":
            window.close()
            event3,values3 = window2.read()
            if event3 == "Save":
                
                sg.popup("New Excel File Created")
                file_name= str(values3['name'])
                new_file = file_name + ".xlsx"
                wb = openpyxl.Workbook()
                path_new = 'C:' + desktop + '\\' + new_file
                wb.save(path_new)
                create = True
                window2.close()
                return path_new

            elif event3 == "Cancel":
                error = 2
                window2.close()
                break
                
            
        elif event2 == "Open":
            open_file = sg.popup_get_file('Open Excel file to start entry...',no_window=True,file_types=(("All Excel Files","*.xlsx"),("All Excel Files","*.xlsm"),("All Excel File","*.xltx"),("All Excel File","*.xltm")))
            window.close()
            if open_file == "":
                error = 2
            return open_file            


def Reset():            # To reset the excel file
    sg.theme("LightBlue2")
    layout2 = [[sg.Text("VERIFY TO RESET THE FILE", font=40)],
            [sg.Text("Password:", size =(10, 1), font=16),sg.InputText(key='pwd', password_char='*', font=16)],
            [sg.Button('Ok'),sg.Button('Cancel')]]

    window = sg.Window("RESET EXCEL", layout2,icon='reset.ico')

    while True:
        event1,values1 = window.read()
        if event1 == "Cancel" or event1 == sg.WIN_CLOSED:
            window.close()
            return False
            
        else:
            if event1 == "Ok":
                if values1['pwd'] == password:
                    sg.popup("Excel file resetted", title='RESET EXCEL',icon='reset.ico')
                    window.close()
                    return True
                   
                elif values1['pwd'] != password:
                    sg.popup("Invalid Password",icon='error.ico')
                    window.close()
                    return False
                    

def delete(sheet):    # For Deleting the rows in the excel
    while(sheet.max_row>1):
        sheet.delete_rows(2)
    book.save(path)

def clear_input():
    for key in values:
        window['Student Initial'].update('')
        window['Student Name'].update('')
        window['Door Number'].update('')
        window['Street Name'].update('')
        window['Area'].update('')
        window['City'].update('')
        window['Pincode'].update('')
        window['Journey to'].update('')
        window['Km'].update('')
        window['Fare'].update('')
    return None

def Check_Empty():      # For checking is there is any empty feild
    for x in val:
        if x == "":
            Empty=False
        else:
            Empty=True
    if Empty == False:
        sg.popup('Empty Feild Not Allowed !',title="Error",icon='error.ico')
    return Empty

def km_fare_check():
    
    if data[12].isdigit() and data[13].isdigit():
        data[13] = int(data[13])*100
        return True
    elif data[12].isdigit() ==False:
        sg.popup("KM Should be a Number",title='Error',icon='error.ico')
        window['Km'].update('')
        return False
    elif data[13].isdigit() ==False:
        sg.popup("Fare Should be a Number",title='Error',icon='error.ico')
        window['Fare'].update('')
        return False
    

def Check_Pincode():    # For Checking Pincode
    B =""
    B = val[9]

    if len(values['Pincode']) != 6:           # Checking whether the pincode is  digits                              
        sg.popup('Pincode must be 6 digits!',title="Invalid Pincode",icon='error.ico')
        window['Pincode'].update('')
        return False
    elif values['Pincode'].isdigit() != True: # Checking whether the pincode is a number
        sg.popup('Pincode must be a number!',title="Invalid Pincode",icon='error.ico')
        window['Pincode'].update('')
        return False
    else:                                    
        if B[:3] == "627" or B[:3] == "628":  # Checking whether the pincode starts with 627 or 628
            return True
        else:
            sg.popup("Pincode should be started with 627xxx or 628xxx",title="Invalid Pincode",icon='error.ico') 
            window['Pincode'].update('')
            return False
   
def Up_Case():                                 # FOR MAKING ALL LETTERS CAPS
    for i in val:                          
        data.append(i.upper())
    


start_up()

while True:
    file = get_file()

    if error == 0:
        path = file #EXCEL FILE NAME
        book = openpyxl.load_workbook(path)
        sheet = book.active
        if create:
            sheet.append(header)
            book.save(path)
            create = False
        break
    elif error == 1:
        sg.popup_auto_close("No Excel File Selected! Closing The Software...",title='Error',auto_close_duration=3,icon='error.ico')
        break
    else:
        continue
        
window = sg.Window("BUSPASS DATA ENTRY", Main_layout,icon='icon.ico') # Creating the layout of the window
password = "jhss"

while True: # Loop
    if error == 1: break

    data = []
    Empty = False

    event, values = window.read()        # Reading the input as values and button pressed as event
    
    if event == sg.WIN_CLOSED or event == 'Exit':
        break
    
    elif event == 'Clear':
        clear_input()

    elif event == 'Reset':
        if(Reset()):
            delete(sheet)
        
    elif event == 'Submit':
        
        values['Journey from'] = values['Area'] 
        val = list(values.values()) 
        val[1] = (val[1] + "." + val[0])        # JOINING NAME WITH INITIAL
        val.pop(0)                              # REMOVING THE INITIAL FROM THE LIST
        val.insert(0, "St. John's Higher Secondary School")
        val.insert(1, "0071")
        Up_Case()

        if (Check_Pincode() and Check_Empty() and km_fare_check()):
            sheet.append(data)
            book.save(path)
            sg.popup('Data Saved!')
            clear_input()

window.close()